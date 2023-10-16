import openpyxl
import time
import os
from tabulate import tabulate
import pandas as pd

# Função para cadastrar uma nova ocorrência
def novaOcorrencia():
    os.system('cls')
    workbook = openpyxl.load_workbook(r"C:\Users\dev-sistemas-manha\Documents\Visual Studio 2022\vscode\prova\prova.xlsx")
    sheet = workbook["ocorrencias"]
    nome = input("Nome do solicitante: ")
    cpf = int(input("CPF do solicitante (SOMENTE NÚMEROS): "))
    endereco = input("Endereço do solicitante: ")
    tel  = int(input("Telefone do solicitante (SOMENTE NÚMEROS): "))
    hora = input("Horário da ocorrência: ")
    desc = input("Descrição do ocorrido: ")
    obs = input("Observações: ")
    sheet.append([nome, cpf, endereco, tel, hora, desc, obs])
    workbook.save(r"C:\Users\dev-sistemas-manha\Documents\Visual Studio 2022\vscode\prova\prova.xlsx")
    print("Solicitação Cadastrada!")
    time.sleep(3)
    os.system('cls')

# Função para atualizar uma ocorrência existente
def attOcorrencia():
    os.system('cls')
    workbook = openpyxl.load_workbook(r"C:\Users\dev-sistemas-manha\Documents\Visual Studio 2022\vscode\prova\prova.xlsx")
    sheet = workbook["ocorrencias"]
    ocorrencias = []
    workbook = openpyxl.load_workbook(r"C:\Users\dev-sistemas-manha\Documents\Visual Studio 2022\vscode\prova\prova.xlsx")
    sheet = workbook["ocorrencias"]
    for row in sheet.iter_rows(min_row=1, values_only=True):
        ocorrencias.append(row)
    print(tabulate(ocorrencias, headers=[], tablefmt= "fancy_grid"))
    cpf = int(input("Digite o CPF do solicitante para atualizar a ocorrência:"))
    linha = None
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if (int(cell.value) == cpf):
                linha = cell
                break
    if linha == None:
        print("CPF não encontrado!")
        time.sleep(3)
        os.system('cls')
    else:
        obsn = input("Atualizar Observações: ")
        sheet.cell(linha.row, 7, obsn)
        workbook.save(r"C:\Users\dev-sistemas-manha\Documents\Visual Studio 2022\vscode\prova\prova.xlsx")
        print("Solicitação Atualizada Com Sucesso!")
        time.sleep(3)
        os.system('cls')


# Função para listar todos as ocorrencias
def listarOcorrencias():
    os.system('cls')
    ocorrencias = []
    op = "1"
    while op == "1":
        workbook = openpyxl.load_workbook(r"C:\Users\dev-sistemas-manha\Documents\Visual Studio 2022\vscode\prova\prova.xlsx")
        sheet = workbook["ocorrencias"]
        for row in sheet.iter_rows(min_row=1, values_only=True):
            ocorrencias.append(row)
        print(tabulate(ocorrencias, headers=[], tablefmt= "fancy_grid"))
        op = input("Pressione ENTER p/ voltar ao menu: ")
        os.system('cls')

# Menu do sistema:
os.system('cls')
while True:
    print(" SISTEMA DEFESA CIVIL ")
    print("1. Cadastrar Ocorrência ")
    print("2. Atualizar Ocorrências")
    print("3. Listar Ocorrências   ")
    print("0. Sair")
    menu = int(input("Opção: "))

    if(menu==1):
        novaOcorrencia()
    if(menu==2):
        attOcorrencia()
    if(menu==3):
        listarOcorrencias()
    if(menu==0):
        raise SystemExit
